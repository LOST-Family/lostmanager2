#!/usr/bin/env python3
"""Liest die historischen CWL-Monatsblätter aus F2P-CWL.xlsx nach f2pcwl_player_season.

Warum überhaupt: die Clash-API kennt keine CWL-Historie. Alles vor September 2026
existiert ausschließlich in dieser Tabelle, und ohne diese Historie kann der
Einteilungsvorschlag nichts sortieren.

Läuft einmalig. Standardmäßig ein Trockenlauf, der nur berichtet - Schreiben nur
mit --apply.

    python tools/import_cwl_excel.py --xlsx F2P-CWL.xlsx            # Trockenlauf
    python tools/import_cwl_excel.py --xlsx F2P-CWL.xlsx --apply    # schreibt

Die Blätter sind über 21 Monate gewachsen, deshalb wird nichts über feste
Positionen gelesen, sondern über die Kopfzeile:

  Team 2-5  Spalten "Kürzel" (Spielertag), "1Ein".."7Ein" (Soll-Sterne des Tages,
            0 = nicht angegriffen) und "1Per".."7Per" (erreichte Sterne)
  Team 1    hat keine Tag-Spalte. Angriff steht in "Tag 1".."Tag 7" (1/0), die
            Sterne in der Spalte unmittelbar rechts daneben - deren Überschrift
            heißt je nach Monat "Sterne" oder "Sterne Tag N", weshalb sie über
            die Position und nicht über den Namen gefunden wird.
"""

import argparse
import collections
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl fehlt:  pip install openpyxl")

MONATE = {
    "januar": 1, "februar": 2, "märz": 3, "maerz": 3, "april": 4, "mai": 5, "juni": 6,
    "juli": 7, "august": 8, "september": 9, "oktober": 10, "november": 11, "dezember": 12,
}

# Auswertungsblätter, keine Monate.
SKIP_SHEETS = {"Hitrate T1 Mar25-Jan26 ", "Älteste"}

# Zwei Blätter beanspruchen denselben Monat; das zweite ist eine Kopie und würde
# die echten Zahlen überschreiben.
SKIP_DUPLICATE = {"Juni 26_2"}


def norm(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def season_key(sheet_name):
    m = re.match(r"\s*([A-Za-zÄÖÜäöüß]+)\s*(\d{2})\s*$", sheet_name)
    if not m:
        return None
    monat = MONATE.get(norm(m.group(1)))
    if not monat:
        return None
    return f"20{m.group(2)}-{monat:02d}"


def name_variants(raw):
    """Kandidaten für einen Spielernamen.

    In Team 1 steht der Account teils vor, teils in der Klammer:
    "x-IKurama (Relaxo)" gegen "Relaxo (x-lmatatabi)". Deshalb beides probieren.
    """
    n = norm(raw)
    out = [n]
    aussen = re.sub(r"\s*\(.*?\)\s*", " ", n).strip()
    if aussen and aussen != n:
        out.append(aussen)
    for innen in re.findall(r"\(([^)]*)\)", n):
        innen = innen.strip()
        if innen:
            out.append(innen)
    return out


class Resolver:
    """Ordnet einen Anzeigenamen einem Spielertag zu.

    Reihenfolge: erst die im Blatt selbst gefundenen Tag/Name-Paare (die sind
    belegt, nicht geraten), dann die F2P-Mitglieder, zuletzt alle Spieler.

    Heikel ist Team 1: dort steht kein Tag, und es waren immer wieder Gäste aus
    anderen Clans dabei. Ein Gast, der zufällig heißt wie ein F2P-Mitglied, würde
    diesem seine Zahlen unterschieben. Deshalb wird festgehalten, ob eine
    Zuordnung belegt (exakter Name oder gelerntes Tag/Name-Paar) oder nur
    geraten war - Letzteres wird zur Sichtprüfung ausgewiesen.
    """

    def __init__(self, f2p_names, all_names):
        self.learned = {}
        self.f2p = f2p_names
        self.all = all_names
        self.unresolved = collections.Counter()
        self.loose = collections.Counter()

    def learn(self, name, tag):
        self.learned.setdefault(norm(name), set()).add(tag)

    def resolve(self, raw):
        exact = norm(raw)
        for src in (self.learned, self.f2p, self.all):
            hit = src.get(exact)
            if hit and len(hit) == 1:
                return next(iter(hit)), True

        # Über die Klammer wird bewusst NICHT aufgelöst. Die Schreibweise
        # "Account (Person)" ist nicht durchgehalten - bei "x-IKurama (Relaxo)"
        # steht der Account außen, bei "Relaxo (x-lmatatabi)" innen. Wer raten
        # will, bucht früher oder später einen Zweitaccount auf den Haupttag und
        # verschmilzt zwei Spieler. Lieber melden als verfälschen; es geht dabei
        # ohnehin fast nur um Gäste.
        for cand in name_variants(raw)[1:]:
            for src in (self.learned, self.f2p, self.all):
                if cand in src:
                    self.loose[f"{norm(raw)}  (Klammer-Treffer, nicht uebernommen)"] += 1
                    self.unresolved[exact] += 1
                    return None, False
        self.unresolved[exact] += 1
        return None, False


def load_names(path):
    out = {}
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            if "|" not in line:
                continue
            tag, name = line.rstrip("\n").split("|", 1)
            out.setdefault(norm(name), set()).add(tag.strip().upper())
    return out


def header_map(ws, row):
    return {
        str(ws.cell(row=row, column=c).value or "").strip(): c
        for c in range(1, ws.max_column + 1)
    }


GUEST_RE = re.compile("gast", re.IGNORECASE)


def read_block(ws, header_row, first_col_name, tag_col):
    """Liefert je Zeile (anzeigename, tag_oder_None, tage, ist_gast)."""
    hdr = header_map(ws, header_row)
    note_col = hdr.get("Bemerkung") or hdr.get("Hinweis")
    tage = []
    for day in range(1, 8):
        if f"{day}Ein" in hdr:                       # Team 2-5
            per = hdr.get(f"{day}Per")
            if per:
                tage.append((hdr[f"{day}Ein"], per))
        elif f"Tag {day}" in hdr:                    # Team 1
            col = hdr[f"Tag {day}"]
            # Die Sternspalte heißt je nach Monat anders - Position statt Name.
            tage.append((col, col + 1))
    if len(tage) != 7:
        return []

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=first_col_name).value
        if name is None or not str(name).strip():
            break
        tag = None
        if tag_col:
            raw = ws.cell(row=r, column=tag_col).value
            if raw and str(raw).strip().startswith("#"):
                tag = str(raw).strip().upper()
        tage_werte = []
        for ein_c, per_c in tage:
            ein = ws.cell(row=r, column=ein_c).value
            per = ws.cell(row=r, column=per_c).value
            tage_werte.append((
                ein if isinstance(ein, (int, float)) else 0,
                per if isinstance(per, (int, float)) else 0,
            ))
        note = ws.cell(row=r, column=note_col).value if note_col else None
        is_guest = bool(note and GUEST_RE.search(str(note)))
        rows.append((str(name).strip(), tag, tage_werte, is_guest))
    return rows


def parse_workbook(path, resolver):
    wb = openpyxl.load_workbook(path, data_only=True)
    records = {}
    skipped = []
    guests = collections.Counter()

    # Erster Durchgang: Tag/Name-Paare einsammeln, damit Team 1 davon profitiert.
    for ws in wb.worksheets:
        if ws.title in SKIP_SHEETS:
            continue
        for r in range(1, ws.max_row + 1):
            if not re.fullmatch(r"Team\s*\d+", str(ws.cell(row=r, column=1).value or "").strip()):
                continue
            hdr = header_map(ws, r + 1)
            cn = hdr.get("Spieler") or hdr.get("Spieler ")
            ct = hdr.get("Kürzel")
            if not (cn and ct):
                continue
            for name, tag, _tage, _guest in read_block(ws, r + 1, cn, ct):
                if tag:
                    resolver.learn(name, tag)

    for ws in wb.worksheets:
        if ws.title in SKIP_SHEETS:
            continue
        if ws.title in SKIP_DUPLICATE:
            skipped.append((ws.title, "Doppelblatt zum selben Monat"))
            continue
        season = season_key(ws.title)
        if not season:
            skipped.append((ws.title, "Monat nicht erkennbar"))
            continue

        for r in range(1, ws.max_row + 1):
            marker = str(ws.cell(row=r, column=1).value or "").strip()
            if not re.fullmatch(r"Team\s*\d+", marker):
                continue
            team_no = int(re.findall(r"\d+", marker)[0])
            hdr = header_map(ws, r + 1)
            cn = hdr.get("Spieler") or hdr.get("Spieler ")
            if not cn:
                skipped.append((f"{ws.title} / {marker}", "keine Kopfzeile"))
                continue

            for name, tag, tage, is_guest in read_block(ws, r + 1, cn, hdr.get("Kürzel")):
                if is_guest:
                    # Gäste aus anderen Clans spielen ad hoc mit; ihre Zahlen
                    # gehören nicht in die F2P-Historie.
                    guests[norm(name)] += 1
                    continue
                player = tag if tag else resolver.resolve(name)[0]
                if not player:
                    continue
                angriffe = sum(1 for ein, _ in tage if ein and ein > 0)
                sterne = sum(per for ein, per in tage if ein and ein > 0)
                # "raus" zählt Fehltage in halben Schritten: ein ausgelassener
                # Angriff 1,0, ein Angriff ohne Stern 0,5. Gegen 387 Zeilen aus
                # sieben Monaten geprüft - die einzigen Abweichungen sind Spieler,
                # die gar nicht angetreten sind und unten ohnehin wegfallen.
                nicht_angegriffen = sum(1 for ein, _ in tage if not ein or ein <= 0)
                ohne_stern = sum(1 for ein, per in tage if ein and ein > 0 and per <= 0)
                verpasst = nicht_angegriffen * 1.0 + ohne_stern * 0.5
                if angriffe == 0:
                    continue  # gar nicht dabei gewesen
                key = (player, season)
                # Wechselt jemand mitten im Monat das Team, gilt der letzte Block.
                records[key] = {
                    "team_no": team_no,
                    "attacks": angriffe,
                    "stars": sterne,
                    "hitrate": round(sterne / angriffe, 3) if angriffe else None,
                    "days_missed": float(verpasst),
                    "bonus_eligible": sterne >= 8,
                }
    return records, skipped, guests


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--f2p-names", default="f2p_players.txt",
                    help="Tag|Name je Zeile, nur F2P-Mitglieder")
    ap.add_argument("--all-names", default="players.txt",
                    help="Tag|Name je Zeile, alle Spieler")
    ap.add_argument("--apply", action="store_true", help="schreibt SQL nach stdout statt nur zu berichten")
    args = ap.parse_args()

    resolver = Resolver(load_names(args.f2p_names), load_names(args.all_names))
    records, skipped, guests = parse_workbook(args.xlsx, resolver)

    seasons = sorted({s for _, s in records})
    print(f"Datensaetze : {len(records)}", file=sys.stderr)
    print(f"Saisons     : {len(seasons)}  ({seasons[0]} .. {seasons[-1]})", file=sys.stderr)
    print(f"Spieler     : {len({p for p, _ in records})}", file=sys.stderr)
    if skipped:
        print("\nUebersprungen:", file=sys.stderr)
        for what, why in skipped:
            print(f"  {what}: {why}", file=sys.stderr)
    if guests:
        print(f"\nAls Gast markiert, uebersprungen ({sum(guests.values())} Zeilen):", file=sys.stderr)
        for nm, cnt in guests.most_common():
            print(f"  {cnt:>3}x  {nm}", file=sys.stderr)
    if resolver.unresolved:
        print(f"\nNicht zuordenbar - vermutlich weitere Gaeste ({len(resolver.unresolved)}):", file=sys.stderr)
        for nm, cnt in resolver.unresolved.most_common():
            print(f"  {cnt:>3}x  {nm}", file=sys.stderr)
    if resolver.loose:
        print(f"\nKlammer-Namen, bewusst nicht zugeordnet ({len(resolver.loose)}):",
              file=sys.stderr)
        for nm, cnt in resolver.loose.most_common():
            print(f"  {cnt:>3}x  {nm}", file=sys.stderr)

    if not args.apply:
        print("\nTrockenlauf - nichts geschrieben. Mit --apply SQL erzeugen.", file=sys.stderr)
        return

    def q(v):
        if v is None:
            return "NULL"
        if isinstance(v, bool):
            return "TRUE" if v else "FALSE"
        if isinstance(v, (int, float)):
            return str(v)
        return "'" + str(v).replace("'", "''") + "'"

    print("BEGIN;")
    for s in seasons:
        print(f"INSERT INTO f2pcwl_seasons (season, state) VALUES ({q(s)}, 'ABGESCHLOSSEN') "
              f"ON CONFLICT (season) DO NOTHING;")
    for (player, season), d in sorted(records.items()):
        print(
            "INSERT INTO f2pcwl_player_season "
            "(player_tag, season, team_no, attacks, stars, hitrate, days_missed, bonus_eligible) VALUES ("
            f"{q(player)}, {q(season)}, {q(d['team_no'])}, {q(d['attacks'])}, {q(d['stars'])}, "
            f"{q(d['hitrate'])}, {q(d['days_missed'])}, {q(d['bonus_eligible'])}) "
            "ON CONFLICT (player_tag, season) DO UPDATE SET "
            "team_no = EXCLUDED.team_no, attacks = EXCLUDED.attacks, stars = EXCLUDED.stars, "
            "hitrate = EXCLUDED.hitrate, days_missed = EXCLUDED.days_missed, "
            "bonus_eligible = EXCLUDED.bonus_eligible;")
    print("COMMIT;")


if __name__ == "__main__":
    main()
